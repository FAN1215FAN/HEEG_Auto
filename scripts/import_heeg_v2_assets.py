from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
ASSET_ROOT = PROJECT_ROOT / 'src' / 'heeg_auto' / 'v2' / 'assets'
DOCS_ROOT = PROJECT_ROOT / 'docs'

WINDOW_ID_MAP = {
    '患者主界面': 'main.patient_home',
    '创建患者': 'dialog.create_patient',
    '创建检查': 'dialog.create_exam',
    '模拟器确认窗口': 'dialog.simulator_confirm',
    '设备设置': 'dialog.device_settings',
    '实时采集窗口': 'window.realtime_acquisition',
    '阻抗窗口': 'window.impedance',
}

CONTROL_TYPE_ACTION_HINT = {
    'Button': '默认单击',
    'ComboBox': '默认下拉选择',
    'Edit': '默认输入',
    'Text': '默认断言文本',
    'RadioButton': '默认选择单选',
    'Window': '默认等待窗口',
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    if text == 'Property does not exist':
        return ''
    return text


def bool_text(value: Any, default: str = '是') -> str:
    text = normalize_text(value)
    return text or default


def safe_identifier(text: str) -> str:
    cleaned = ''.join(char.lower() if char.isalnum() else '_' for char in text)
    while '__' in cleaned:
        cleaned = cleaned.replace('__', '_')
    return cleaned.strip('_') or 'asset'


def unicode_slug(text: str) -> str:
    chunks = []
    for char in text:
        if char.isascii() and char.isalnum():
            chunks.append(char.lower())
        elif char.isspace():
            chunks.append('_')
        else:
            chunks.append(f'u{ord(char):x}')
    slug = ''.join(chunks)
    while '__' in slug:
        slug = slug.replace('__', '_')
    return slug.strip('_') or 'asset'


def build_window_assets(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, data_only=True)
    window_assets: list[dict[str, Any]] = []
    element_assets: list[dict[str, Any]] = []
    missing_automation: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        headers = [normalize_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_index = {header: idx for idx, header in enumerate(headers)}
        sheet_window_id = WINDOW_ID_MAP.get(ws.title, f'window.{unicode_slug(ws.title)}')
        sheet_window_label = ws.title
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = {header: row[idx] if idx < len(row) else None for header, idx in header_index.items()}
            asset_type = normalize_text(values.get('资产类型'))
            label = normalize_text(values.get('中文名称'))
            owner_window = normalize_text(values.get('所属窗口'))
            automation_id = normalize_text(values.get('AutomationId'))
            control_type = normalize_text(values.get('ControlType'))
            class_name = normalize_text(values.get('ClassName'))
            name = normalize_text(values.get('Name'))
            unique = bool_text(values.get('是否唯一'))
            anchor = normalize_text(values.get('锚点'))
            anchor_values = ['是'] if anchor == '是' else []
            description = CONTROL_TYPE_ACTION_HINT.get(control_type, '')
            if asset_type == '窗口':
                window_assets.append(
                    {
                        '窗口标识': sheet_window_id,
                        '资产类型': '窗口',
                        '中文名称': sheet_window_label,
                        '所属窗口': owner_window or sheet_window_label,
                        'AutomationId': automation_id,
                        'ControlType': control_type or 'Window',
                        'Name': name or owner_window or sheet_window_label,
                        'ClassName': class_name or 'Window',
                        '是否唯一': unique,
                        '锚点元素': anchor_values,
                        '用途说明': f'{sheet_window_label}窗口',
                    }
                )
                continue
            if not asset_type.startswith('元素'):
                continue
            if automation_id:
                element_key = safe_identifier(automation_id)
            else:
                element_key = unicode_slug(label or control_type or 'asset')
                missing_automation.append(
                    {
                        'sheet': ws.title,
                        'label': label,
                        'control_type': control_type,
                        'name': name,
                    }
                )
            element_assets.append(
                {
                    '元素标识': f'{sheet_window_id}.{element_key}',
                    '资产类型': asset_type,
                    '中文名称': label,
                    '所属窗口': sheet_window_label,
                    'AutomationId': automation_id,
                    'ControlType': control_type,
                    'Name': name,
                    'ClassName': class_name,
                    '是否唯一': unique,
                    '锚点元素': anchor_values,
                    '用途说明': description,
                }
            )
    return window_assets, element_assets, missing_automation


def dump_yaml(path: Path, top_key: str, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        yaml.safe_dump({top_key: rows}, allow_unicode=True, sort_keys=False, width=120),
        encoding='utf-8',
    )


def write_docs(window_assets: list[dict[str, Any]], element_assets: list[dict[str, Any]], missing_automation: list[dict[str, Any]]) -> None:
    docs_path = DOCS_ROOT / 'V2资产总表.md'
    lines = [
        '# V2资产总表',
        '',
        f'- 窗口资产数：{len(window_assets)}',
        f'- 元素资产数：{len(element_assets)}',
        f'- 缺少 AutomationId 的元素数：{len(missing_automation)}',
        '',
        '## 窗口资产',
        '',
        '| 窗口标识 | 中文名称 | Name | ControlType |',
        '| --- | --- | --- | --- |',
    ]
    for item in window_assets:
        lines.append(f"| {item['窗口标识']} | {item['中文名称']} | {item.get('Name','')} | {item.get('ControlType','')} |")
    lines.extend(['', '## 元素资产', ''])
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in element_assets:
        grouped.setdefault(item['所属窗口'], []).append(item)
    for window, rows in grouped.items():
        lines.extend([
            f'### {window}',
            '',
            '| 元素标识 | 中文名称 | AutomationId | ControlType | 是否唯一 |',
            '| --- | --- | --- | --- | --- |',
        ])
        for item in rows:
            lines.append(
                f"| {item['元素标识']} | {item['中文名称']} | {item.get('AutomationId','')} | {item.get('ControlType','')} | {item.get('是否唯一','')} |"
            )
        lines.append('')
    docs_path.write_text('\n'.join(lines), encoding='utf-8')

    gap_path = DOCS_ROOT / 'V2资产缺口清单.md'
    gap_lines = [
        '# V2资产缺口清单',
        '',
        '以下元素缺少 `AutomationId`，建议优先让研发补齐：',
        '',
        '| 所属窗口 | 中文名称 | ControlType | Name |',
        '| --- | --- | --- | --- |',
    ]
    for item in missing_automation:
        gap_lines.append(f"| {item['sheet']} | {item['label']} | {item['control_type']} | {item['name']} |")
    if not missing_automation:
        gap_lines.append('| - | - | - | - |')
    gap_path.write_text('\n'.join(gap_lines), encoding='utf-8')


def main() -> int:
    parser = argparse.ArgumentParser(description='Import HEEG Excel label workbook into V2 YAML assets.')
    parser.add_argument('workbook', type=Path, help='Path to the Excel workbook.')
    args = parser.parse_args()

    window_assets, element_assets, missing_automation = build_window_assets(args.workbook)
    dump_yaml(ASSET_ROOT / 'windows' / 'heeg_windows.yaml', '窗口资产', window_assets)
    dump_yaml(ASSET_ROOT / 'elements' / 'heeg_elements.yaml', '元素资产', element_assets)
    write_docs(window_assets, element_assets, missing_automation)

    print(f'Imported windows: {len(window_assets)}')
    print(f'Imported elements: {len(element_assets)}')
    print(f'Missing automation ids: {len(missing_automation)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
